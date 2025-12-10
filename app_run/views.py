from django.shortcuts import render, get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.views import APIView
from django.conf import settings
from django.contrib.auth.models import User
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from openpyxl import load_workbook
from .serializers import (RunSerializer, UserSerializer, AthleteInfoSerializer, ChallengeSerializer, PositionSerializer,
                          CollectibleItemSerializer, UserItemsSerializer)
from .models import Run, AthleteInfo, Challenge, Position, CollectibleItem
from . import utils, enum


@api_view(['GET'])
def company_details_view(request):
    return Response({
        'company_name': settings.COMPANY_NAME,
        'slogan': settings.SLOGAN,
        'contacts': settings.CONTACTS,
    })


class Pagination(PageNumberPagination):
    # page_size = 3  # Количество объектов на странице по умолчанию (не обязательный параметр)
    page_size_query_param = 'size'  # Разрешаем изменять количество объектов через query параметр size в url
    max_page_size = 50  # Ограничиваем максимальное количество объектов на странице


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.all().select_related('athlete')
    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']
    ordering = ['id']
    pagination_class = Pagination


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.filter(is_superuser=False)
    serializer_class = UserSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']
    ordering = ['id']
    pagination_class = Pagination

    #Если выполняется retrieve(GET+ID) вызываем другой сериализатор
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return UserItemsSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = self.queryset
        user_type = self.request.query_params.get('type', None)
        if user_type == 'coach':
            qs = qs.filter(is_staff=True)
        elif user_type == 'athlete':
            qs = qs.filter(is_staff=False)
        #Можно оставить, можно убрать, на производительность при retrive особо не повлияет, в любом случае два запроса.
        if self.action == "retrieve":
            qs = qs.prefetch_related("items")
        return qs


class RunStartView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == Run.Status.INIT:
            run.status = Run.Status.IN_PROGRESS
            run.save()
            utils.challenge_check(enum.ChallengeEvent.RUN_STARTED, run)
            return Response({'detail': f'Статус объекта {run_id} обновлен на {Run.Status.IN_PROGRESS}.'})
        else:
            return Response({'detail': f'Статус объекта {run_id} - {run.status}. Обновление статуса не выполнено.'},
                            status=status.HTTP_400_BAD_REQUEST)


class RunStopView(APIView):
    def patch(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == Run.Status.IN_PROGRESS:
            run.status = Run.Status.FINISHED
            run.distance = run.distance_calculation()
            run.save()
            utils.challenge_check(enum.ChallengeEvent.RUN_FINISHED, run)
            run_positions = Position.objects.filter(run=run).order_by('date_time')
            if len(run_positions) > 1:
                run.run_time_seconds = (
                    int((run_positions[len(run_positions)-1].date_time - run_positions[0].date_time).total_seconds()))
            else:
                run.run_time_seconds = 0
            run.save()
            return Response({'detail': f'Статус объекта {run_id} обновлен на {Run.Status.FINISHED}.',
                             'distance': run.distance})
        else:
            return Response({'detail': f'Статус объекта {run_id} - {run.status}. Обновление статуса не выполнено.'},
                            status=status.HTTP_400_BAD_REQUEST)

    def post(self, request, run_id):
        return self.patch(request, run_id)


class AthleteInfoView(APIView):
    user = None

    #вызывается перед вызовом любого другого описанного вида вызовов
    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        self.user = get_object_or_404(User, id=kwargs['user_id'])

    def get(self, request, user_id):
        athlete_info, created = AthleteInfo.objects.get_or_create(user=self.user, defaults={'user': self.user, 'goals': '',
                                                                                          'weight': 0})
        serializer = AthleteInfoSerializer(athlete_info)
        return Response(serializer.data)

    def put(self, request, user_id):
        athlete_info, created = AthleteInfo.objects.get_or_create(user=self.user, defaults={'user': self.user,
                                                                            'goals': request.data.get('goals', ''),
                                                                            'weight': request.data.get('weight', 0)})
        serializer = AthleteInfoSerializer(athlete_info, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ChallengeViewSet(viewsets.ModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['athlete']


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    filter_backends = [DjangoFilterBackend]
    search_fields = ['run']

    # метод вызывается при вызове POST
    def perform_create(self, serializer):
        # Стандартный функционал
        position = serializer.save()
        # Дополнительная логика:
        position.collect_items()


class CollectibleItemViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class UploadFileView(APIView):
    def post(self, request):
        excel_file = request.FILES.get('file')
        if excel_file is None:
            pass
        if not excel_file.name.endswith((".xlsx", ".xls")):
            pass
        try:
            workbook = load_workbook(excel_file, data_only=True)

            # первая страница
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            headers = list(rows[0])
            incorrect_headers_name = headers.index('URL')
            headers[incorrect_headers_name] = 'picture'
            data_rows = rows[1:]
            result_rows = []
            incorrect_rows = []
            for row in data_rows:
                row_dict = {header.lower(): value for header, value in zip(headers, row)}
                result_rows.append(row_dict)
            for row in result_rows:
                serializer = CollectibleItemSerializer(data=row, partial=True)
                if not serializer.is_valid(raise_exception=False):
                    incorrect_rows.append([v for v in row.values()])
                else:
                    serializer.save()
            return Response(incorrect_rows)
        except Exception as e:
            return Response(f'Загрузка документа не выполнена, ошибка: {e}', status=status.HTTP_400_BAD_REQUEST)
